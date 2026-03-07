"""
Excel import for CSR Consolidated Report: parse .xlsx and create plans + activities (+ optional realizations).

Expected sheet: first row = headers (case-insensitive, trimmed). Column names are matched flexibly.
Supported column names (examples):
  - Site: "Site", "Site Code", "Code", "Code site"
  - Year: "Year", "Année", "Annee"
  - Category: "Category", "Catégorie", "Categorie"
  - Activity: "Activity Number", "No", "N°", "Numéro", "Activity No"
  - Title: "Title", "Titre", "Activity Title"
  - Description: "Description"
  - Planned Budget: "Planned Budget", "Budget prévu", "Budget"
  - Organization: "Organization", "Organisation"
  - Contract Type: "Contract Type", "Type contrat"
  - Realized Budget: "Realized Budget", "Budget réalisé"
  - Participants: "Participants", "Participant"
  - Realization Year/Month: "Realization Year", "Year" (realization), "Month"
"""
import re
from decimal import Decimal
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _norm(s: str) -> str:
    """Normalize header: lowercase, strip, collapse spaces and newlines."""
    if s is None or not isinstance(s, str):
        return ""
    return " ".join(str(s).replace("\n", " ").lower().strip().split())


# Header variants -> internal key. Order matches "2024 CSR Consolidated Report Form":
# Activity N, Region, country, Plant, Activity Title/ description, Category, Nature of collaboration,
# Start year, Edition, Nbr of internal Participants, Total HC, ..., Estimated Budget in €, Actual Budget in €,
# Action Impact In Numbers, Action Impact Unit, Organizer, External Partner, Number of External Partners
HEADER_MAP = {
    "site": ["plant", "site", "site code", "code", "code site"],
    "year": ["start year", "year", "année", "annee", "an"],
    "category": ["category", "catégorie", "categorie", "categorie csr"],
    "activity_number": ["activity n", "activity number", "no", "n°", "numéro", "numero", "activity no"],
    "title": ["activity title/ description", "activity title", "title", "titre", "activity", "intitulé", "intitule", "nom"],
    "description": ["description", "desc"],
    "planned_budget": ["estimated budget in €", "estimated budget", "planned budget", "budget prévu", "budget prevu", "budget"],
    "organization": ["organization", "organisation", "org"],
    "contract_type": ["contract type", "type contrat", "type de contrat"],
    "realized_budget": ["actual budget in €", "actual budget", "realized budget", "budget réalisé", "budget reel"],
    "participants": ["nbr of internal participants", "participants", "participant", "nombre participants", "nb participants"],
    "impact_actual": ["action impact in numbers", "impact actual", "impact réalisé", "action impact actual"],
    "impact_unit": ["action impact unit", "action impact\nunit", "impact unit"],
    "realization_year": ["realization year", "year realization", "année réalisation"],
    "realization_month": ["realization month", "month", "mois", "month realization"],
    "volunteer_hours": ["volunteer hours", "heures volontariat", "heures volontaires"],
    "collaboration_nature": ["nature of collaboration", "nature collaboration"],
    "organizer": ["organizer", "organisateur"],
    "number_external_partners": ["number of external partners", "number of external partners_by", "nb external partners"],
}


def _find_column_index(headers: list[str]) -> dict[str, int]:
    """Return dict mapping internal key -> 0-based column index (only if header matches)."""
    normalized = [_norm(h) for h in headers]
    result = {}
    for key, variants in HEADER_MAP.items():
        for v in variants:
            for i, h in enumerate(normalized):
                if v in h or h in v or h == v:
                    result[key] = i
                    break
            if key in result:
                break
    return result


def _cell_value(ws: Worksheet, row: int, col: int) -> Any:
    """Get cell value (1-based row/col in openpyxl)."""
    try:
        return ws.cell(row=row, column=col + 1).value
    except Exception:
        return None


def _safe_int(val: Any) -> Optional[int]:
    if val is None or val == "":
        return None
    if isinstance(val, int):
        return val
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


def _safe_float(val: Any) -> Optional[float]:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, Decimal):
        return float(val)
    s = str(val).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(re.sub(r"[^\d.\-]", "", s))
    except (TypeError, ValueError):
        return None


def _safe_str(val: Any, max_len: int = 255) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    return s[:max_len] if max_len else s


def parse_excel_rows(file_path: str) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Parse Excel file and return list of row dicts (keys = internal keys) and list of parse errors.
    Does not touch the database.
    """
    errors = []
    rows = []
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            errors.append("Feuille active vide")
            return rows, errors
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row < 2 or max_col < 2:
            errors.append("Le fichier doit contenir une ligne d'en-têtes et au moins une ligne de données")
            return rows, errors

        headers = []
        for c in range(max_col):
            headers.append(_cell_value(ws, 1, c) or "")

        col_map = _find_column_index(headers)
        if not col_map.get("site"):
            try:
                # Fallback: column 4 (1-based) = Plant in standard 2024 form order
                if max_col >= 4 and _norm(headers[3]) in ("plant", "site", "code"):
                    col_map["site"] = 3
            except IndexError:
                pass
        if not col_map.get("site"):
            errors.append("Colonne 'Plant' (ou Site/Code) non trouvée. Vérifiez que la première ligne contient les en-têtes.")

        for r in range(2, max_row + 1):
            row_data = {}
            for key, col_idx in col_map.items():
                val = _cell_value(ws, r, col_idx)
                if val is None:
                    continue
                if isinstance(val, str) and not val.strip():
                    continue
                row_data[key] = val
            # Include only rows that have Plant/site (required). Year/category can be defaulted in the route.
            if row_data.get("site"):
                rows.append(row_data)
        wb.close()
    except Exception as e:
        errors.append(f"Erreur lecture Excel: {e}")
    return rows, errors


def build_import_result(
    rows: list[dict],
    site_id_override: Optional[str] = None,
    year_override: Optional[int] = None,
) -> dict[str, Any]:
    """
    From parsed rows, build a structure suitable for the API to apply to DB.
    Returns dict with: site_id, year, plan_id, activities_created, realized_created, errors, warnings.
    Does not perform DB writes; the route will use this to get (site_id, year, rows) and then create plan/activities.
    """
    return {
        "site_id_override": site_id_override,
        "year_override": year_override,
        "rows": rows,
    }
