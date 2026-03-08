"""
Excel import for CSR Consolidated Report: parse .xlsx and create plans + activities (+ optional realizations).

Columns are read by fixed index (0-based) per "2024 CSR Consolidated Report Form - Group figures":
  0=Activity N, 1=Region, 2=country, 3=Plant, 4=Activity Title/description, 5=Category, 6=Nature of collaboration,
  7=Year, 8=Start year, 9=Edition, 10=Nbr of internal Participants, 11=Total HC, 12=Percentage out of all the employees %,
  13=Estimated Budget in €, 14=Actual Budget in €, 15=Action Impact In Numbers, 16=Action Impact Unit,
  17=Organizer, 18=External Partner, 19=Number of External Partners
"""
import re
from decimal import Decimal
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# Fixed column indices (0-based) for "2024 CSR Consolidated Report Form - Group figures"
COLUMN_INDICES = {
    "activity_number": 0,
    "region": 1,
    "country": 2,
    "site": 3,  # Plant
    "title": 4,  # Activity Title/ description
    "category": 5,
    "collaboration_nature": 6,
    "year": 7,
    "start_year": 8,
    "edition": 9,
    "participants": 10,  # Nbr of internal Participants (used for both planned_volunteers and realized participants)
    "total_hc": 11,
    "percentage_employees": 12,
    "planned_budget": 13,  # Estimated Budget in €
    "realized_budget": 14,  # Actual Budget in €
    "impact_actual": 15,  # Action Impact In Numbers (also impact_target for planned)
    "impact_unit": 16,
    "organizer": 17,
    "external_partner": 18,
    "number_external_partners": 19,
}


def _cell_value(ws: Worksheet, row: int, col: int) -> Any:
    """Get cell value (1-based row/col in openpyxl)."""
    try:
        return ws.cell(row=row, column=col + 1).value
    except Exception:
        return None


def _safe_int(val: Any) -> Optional[int]:
    if val is None or val == "":
        return None
    if isinstance(val, int):
        return val
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


def _safe_float(val: Any) -> Optional[float]:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, Decimal):
        return float(val)
    s = str(val).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(re.sub(r"[^\d.\-]", "", s))
    except (TypeError, ValueError):
        return None


def _safe_str(val: Any, max_len: int = 255) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    return s[:max_len] if max_len else s


def parse_excel_rows(file_path: str) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Parse Excel file and return list of row dicts (keys = internal keys) and list of parse errors.
    Does not touch the database.
    """
    errors = []
    rows = []
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            errors.append("Feuille active vide")
            return rows, errors
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row < 2 or max_col < 2:
            errors.append("Le fichier doit contenir une ligne d'en-têtes et au moins une ligne de données")
            return rows, errors

        if max_col < 4:
            errors.append("Le fichier doit avoir au moins 4 colonnes (Plant = colonne 4).")

        for r in range(2, max_row + 1):
            row_data = {}
            for key, col_idx in COLUMN_INDICES.items():
                if col_idx >= max_col:
                    continue
                val = _cell_value(ws, r, col_idx)
                if val is None:
                    continue
                if isinstance(val, str) and not val.strip():
                    continue
                row_data[key] = val
            # planned_volunteers and impact_target use same columns as participants and impact_actual
            if "participants" in row_data and "planned_volunteers" not in row_data:
                row_data["planned_volunteers"] = row_data["participants"]
            if "impact_actual" in row_data and "impact_target" not in row_data:
                row_data["impact_target"] = row_data["impact_actual"]
            # year: use start_year if year not present (col 8 = Start year)
            if "year" not in row_data and "start_year" in row_data:
                row_data["year"] = row_data["start_year"]
            # Include only rows that have Plant/site (required)
            if row_data.get("site"):
                rows.append(row_data)
        wb.close()
    except Exception as e:
        errors.append(f"Erreur lecture Excel: {e}")
    return rows, errors


def build_import_result(
    rows: list[dict],
    site_id_override: Optional[str] = None,
    year_override: Optional[int] = None,
) -> dict[str, Any]:
    """
    From parsed rows, build a structure suitable for the API to apply to DB.
    Returns dict with: site_id, year, plan_id, activities_created, realized_created, errors, warnings.
    Does not perform DB writes; the route will use this to get (site_id, year, rows) and then create plan/activities.
    """
    return {
        "site_id_override": site_id_override,
        "year_override": year_override,
        "rows": rows,
    }
